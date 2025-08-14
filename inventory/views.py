from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.template.loader import get_template
from django.db.models import Sum, Q
from xhtml2pdf import pisa
import openpyxl
from io import BytesIO
from datetime import datetime

from .models import Product, Movement, Supplier, MovementType, UnitOfMeasurement
from .forms import (
    ProductForm, 
    MovementForm, 
    SupplierForm, 
    ReportForm,
    ImportForm
)

class DashboardView(LoginRequiredMixin, ListView):
    template_name = 'inventory/dashboard.html'
    context_object_name = 'products'
    
    def get_queryset(self):
        return Product.objects.order_by('-current_stock')[:10]
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Productos con bajo stock
        low_stock = Product.objects.filter(current_stock__lte=5)
        context['low_stock'] = low_stock
        
        # Últimos movimientos
        context['recent_movements'] = Movement.objects.order_by('-date')[:5]
        
        # Estadísticas
        context['total_products'] = Product.objects.count()
        context['total_movements'] = Movement.objects.count()
        
        return context

class ProductListView(LoginRequiredMixin, ListView):
    model = Product
    template_name = 'inventory/product_list.html'
    context_object_name = 'products'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.GET.get('search')
        
        if search:
            queryset = queryset.filter(
                Q(code__icontains=search) |
                Q(description__icontains=search)
            )
        
        return queryset.order_by('description')

class ProductCreateView(LoginRequiredMixin, CreateView):
    model = Product
    form_class = ProductForm
    template_name = 'inventory/product_form.html'
    success_url = reverse_lazy('inventory:product_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Producto creado exitosamente.')
        return super().form_valid(form)

class ProductUpdateView(LoginRequiredMixin, UpdateView):
    model = Product
    form_class = ProductForm
    template_name = 'inventory/product_form.html'
    success_url = reverse_lazy('inventory:product_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Producto actualizado exitosamente.')
        return super().form_valid(form)

class ProductDeleteView(LoginRequiredMixin, DeleteView):
    model = Product
    template_name = 'inventory/product_confirm_delete.html'
    success_url = reverse_lazy('inventory:product_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Producto eliminado exitosamente.')
        return super().delete(request, *args, **kwargs)

class MovementListView(LoginRequiredMixin, ListView):
    model = Movement
    template_name = 'inventory/movement_list.html'
    context_object_name = 'movements'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filtros
        product_id = self.request.GET.get('product')
        movement_type = self.request.GET.get('type')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        
        if product_id:
            queryset = queryset.filter(product_id=product_id)
        
        if movement_type:
            queryset = queryset.filter(movement_type_id=movement_type)
        
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        
        return queryset.order_by('-date', '-created_at')

class MovementCreateView(LoginRequiredMixin, CreateView):
    model = Movement
    form_class = MovementForm
    template_name = 'inventory/movement_form.html'
    
    def get_success_url(self):
        return reverse_lazy('inventory:movement_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        response = super().form_valid(form)
        messages.success(self.request, 'Movimiento registrado exitosamente.')
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['products'] = Product.objects.all()
        return context

class MovementDetailView(LoginRequiredMixin, DetailView):
    model = Movement
    template_name = 'inventory/movement_detail.html'
    context_object_name = 'movement'

def generate_report(request):
    if request.method == 'POST':
        form = ReportForm(request.POST)
        if form.is_valid():
            # Obtener datos del formulario
            report_type = form.cleaned_data['report_type']
            date_from = form.cleaned_data['date_from']
            date_to = form.cleaned_data['date_to']
            product = form.cleaned_data['product']
            movement_type = form.cleaned_data['movement_type']
            output_format = form.cleaned_data['output_format']
            
            # Filtrar datos según parámetros
            movements = Movement.objects.all()
            
            if date_from:
                movements = movements.filter(date__gte=date_from)
            
            if date_to:
                movements = movements.filter(date__lte=date_to)
            
            if product:
                movements = movements.filter(product=product)
            
            if movement_type:
                movements = movements.filter(movement_type=movement_type)
            
            # Generar reporte según tipo
            if report_type == 'movements':
                data = movements.order_by('-date')
                template = 'inventory/report_pdf.html'
                context = {
                    'title': 'Reporte de Movimientos',
                    'data': data,
                    'date_from': date_from,
                    'date_to': date_to,
                }
            elif report_type == 'stock':
                products = Product.objects.all().order_by('description')
                if product:
                    products = products.filter(id=product.id)
                
                template = 'inventory/report_pdf.html'
                context = {
                    'title': 'Reporte de Stock',
                    'data': products,
                    'date_from': date_from,
                    'date_to': date_to,
                }
            
            # Generar respuesta según formato
            if output_format == 'pdf':
                response = HttpResponse(content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="reporte_{report_type}_{datetime.now().strftime("%Y%m%d")}.pdf"'
                
                html = get_template(template).render(context)
                pisa_status = pisa.CreatePDF(html, dest=response)
                
                if pisa_status.err:
                    return HttpResponse('Error al generar PDF')
                return response
            
            elif output_format == 'excel':
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="reporte_{report_type}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
                
                wb = openpyxl.Workbook()
                ws = wb.active
                
                if report_type == 'movements':
                    ws.title = "Movimientos"
                    headers = [
                        "Número", "Fecha", "Producto", "Tipo", 
                        "Cantidad", "Proveedor", "Entregado a", 
                        "Precio Unitario", "Precio Total"
                    ]
                    ws.append(headers)
                    
                    for movement in movements:
                        ws.append([
                            movement.movement_number,
                            movement.date.strftime("%Y-%m-%d"),
                            movement.product.description,
                            movement.movement_type.name,
                            movement.quantity,
                            movement.supplier.name if movement.supplier else "",
                            movement.delivered_to,
                            movement.unit_price,
                            movement.total_price
                        ])
                
                elif report_type == 'stock':
                    ws.title = "Stock"
                    headers = [
                        "Código", "Descripción", "Unidad", 
                        "Stock Actual", "Stock Mínimo", 
                        "Precio Unitario", "Precio Total"
                    ]
                    ws.append(headers)
                    
                    for product in products:
                        ws.append([
                            product.code,
                            product.description,
                            product.unit_of_measurement.name,
                            product.current_stock,
                            product.min_stock,
                            product.unit_price,
                            product.total_price
                        ])
                
                wb.save(response)
                return response
            
    else:
        form = ReportForm()
    
    return render(request, 'inventory/report_form.html', {'form': form})

def import_data(request):
    if request.method == 'POST':
        form = ImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            file_type = form.cleaned_data['file_type']
            
            try:
                if file_type == 'products':
                    wb = openpyxl.load_workbook(file)
                    sheet = wb.active
                    
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        code, description, unit_name, stock, unit_price = row
                        
                        unit, _ = UnitOfMeasurement.objects.get_or_create(
                            name=unit_name,
                            defaults={'abbreviation': unit_name[:3].upper()}
                        )
                        
                        Product.objects.update_or_create(
                            code=code,
                            defaults={
                                'description': description,
                                'unit_of_measurement': unit,
                                'current_stock': stock or 0,
                                'unit_price': unit_price or 0
                            }
                        )
                    
                    messages.success(request, 'Productos importados exitosamente.')
                
                elif file_type == 'movements':
                    wb = openpyxl.load_workbook(file)
                    sheet = wb.active
                    
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        date, product_code, type_name, quantity, supplier_name, delivered_to = row
                        
                        product = get_object_or_404(Product, code=product_code)
                        movement_type, _ = MovementType.objects.get_or_create(
                            name=type_name,
                            defaults={'code': type_name[:3].upper(), 'is_entry': 'entrada' in type_name.lower()}
                        )
                        
                        supplier = None
                        if supplier_name:
                            supplier, _ = Supplier.objects.get_or_create(
                                name=supplier_name,
                                defaults={'code': supplier_name[:3].upper()}
                            )
                        
                        Movement.objects.create(
                            date=date,
                            product=product,
                            movement_type=movement_type,
                            quantity=quantity,
                            supplier=supplier,
                            delivered_to=delivered_to,
                            unit_price=product.unit_price
                        )
                    
                    messages.success(request, 'Movimientos importados exitosamente.')
                
                return redirect('inventory:import_export')
            
            except Exception as e:
                messages.error(request, f'Error al importar datos: {str(e)}')
    
    else:
        form = ImportForm()
    
    return render(request, 'inventory/import_export.html', {'form': form})

def export_data(request):
    file_type = request.GET.get('type', 'products')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{file_type}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    if file_type == 'products':
        ws.title = "Productos"
        headers = [
            "Código", "Descripción", "Unidad de Medida", 
            "Stock Actual", "Precio Unitario"
        ]
        ws.append(headers)
        
        for product in Product.objects.all():
            ws.append([
                product.code,
                product.description,
                product.unit_of_measurement.name,
                product.current_stock,
                product.unit_price
            ])
    
    elif file_type == 'movements':
        ws.title = "Movimientos"
        headers = [
            "Fecha", "Producto (Código)", "Tipo de Movimiento", 
            "Cantidad", "Proveedor", "Entregado a"
        ]
        ws.append(headers)
        
        for movement in Movement.objects.all():
            ws.append([
                movement.date.strftime("%Y-%m-%d"),
                f"{movement.product.description} ({movement.product.code})",
                movement.movement_type.name,
                movement.quantity,
                movement.supplier.name if movement.supplier else "",
                movement.delivered_to
            ])
    
    wb.save(response)
    return response